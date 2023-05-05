"""
Dettectinator - The Python library to your DeTT&CT YAML files.
Authors:
    Martijn Veken, Sirius Security
    Ruben Bouman, Sirius Security
License: GPL-3.0 License
"""

from argparse import ArgumentParser
from collections.abc import Iterable
import urllib3
import re
from pypdf import PdfReader

# Disable SSL certificate warnings for dev purposes:
urllib3.disable_warnings()


class GroupBase:
    """
    Base class for importing group/technique/software data
    """

    def __init__(self, parameters: dict) -> None:
        self._parameters = parameters

        self._campaign = self._parameters.get('campaign', '')
        self._group = self._parameters.get('group', '')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('-gr', '--group',
                            help='Name of the group.', default='')
        parser.add_argument('-ca', '--campaign',
                            help='Name of the campaign of the group.', default='')

    def get_attack_groups(self) -> dict:
        """
        Retrieves group/technique/software data from a data source
        :return: Dictionary, example:  {'APT1': {'campaign': 'P0wn them all', 'techniques': ['T1566.002', 'T1059.001',
                                        'T1053.005'], 'software': ['S0002']}}
        """
        groups = {}

        for group, campaign, techniques, software in self.get_data_from_source():
            campaign = campaign or self._campaign
            group = group or self._group

            groups[group] = {'campaign': campaign,
                             'techniques': techniques,
                             'software': software}

        return groups

    def get_data_from_source(self) -> Iterable:
        """
        Retrieves group/technique/software data from the source
        :return: Iterable, yields group, campaign, techniques, software
        """
        raise NotImplementedError()

    @staticmethod
    def _create_unique_sorted_list(items: list) -> list:
        return sorted(list(set(items)))


class GroupExcel(GroupBase):
    """
    Sample plugin to import data from Excel. Each group has its own tab and on that tab
    the techniques are listed.
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('GroupExcel: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        GroupBase.set_plugin_params(parser)

        parser.add_argument('--file', help='Path of the Excel file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Retrieves group/technique/software data from the source
        See examples/import_group.xlsx for an example Excel file
        :return: Iterable, yields group, campaign, techniques, software
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        import openpyxl
        wb = openpyxl.load_workbook(filename=file, data_only=True)

        for sheet in wb.worksheets:
            group = sheet.title
            techniques = []

            for rowNumber in range(2, sheet.max_row + 1):
                technique = sheet.cell(row=rowNumber, column=1).value

                if technique:
                    techniques.append(technique)

            yield group, None, self._create_unique_sorted_list(techniques), []


class GroupPdf(GroupBase):
    """
    Sample plugin to import data from a PDF. It uses a regular expression to fetch techniques and software
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('GroupPdf: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        GroupBase.set_plugin_params(parser)

        parser.add_argument('--file', help='Path of the Excel file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Retrieves group/technique/software data from the source
        :return: Iterable, yields group, campaign, techniques, software
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        reader = PdfReader(file)

        text = ''
        for page in reader.pages:
            text += f'{page.extract_text()} '

        # Get ATT&CK Technique ID's from the text
        pattern = r'T[0-9]{4}\.[0-9]{3}|T[0-9]{4}'
        techniques = self._create_unique_sorted_list(re.findall(pattern, text))

        # Get ATT&CK Software ID's from the text
        pattern = r'S[0-9]{4}'
        software = self._create_unique_sorted_list(re.findall(pattern, text))

        yield None, None, techniques, software
