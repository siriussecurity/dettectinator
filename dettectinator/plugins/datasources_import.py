"""
Dettectinator - The Python library to your DeTT&CT YAML files.
Authors:
    Martijn Veken, Sirius Security
    Ruben Bouman, Sirius Security
License: GPL-3.0 License
"""

import json
import xml.etree.ElementTree as ElementTree
import requests
import re
import sys
import os

from argparse import ArgumentParser
from collections.abc import Iterable
from xml.etree.ElementTree import Element

try:
    # When dettectinator is installed as python library
    from dettectinator.plugins.support.authentication import Azure
except ModuleNotFoundError:
    # When dettectinator is not installed as python library
    sys.path.append(os.path.dirname(os.path.abspath(__file__).replace('plugins', '')))
    from plugins.support.authentication import Azure


class DatasourceBase:
    """
    Base class for importing datasource/product data
    """

    def __init__(self, parameters: dict) -> None:
        self._parameters = parameters

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        raise NotImplementedError()

    def get_attack_datasources(self, default_applicable_to: list) -> dict:
        """
        Retrieves datasource/product data from a data source
        :param default_applicable_to: Systems that the data sources are applicable to.
        :return: Dictionary, example: {"User Account Creation":[{"applicable_to":["test"],"available_for_data_analytics":true,"products":["DeviceEvents: UserAccountCreated"]}]}
        """

        data_sources = {}

        for datasource, product, applicable_to in self.get_data_from_source():
            applicable_to = applicable_to or default_applicable_to

            if datasource not in data_sources.keys():
                record = {'applicable_to': applicable_to, 'available_for_data_analytics': True, 'products': []}
                data_sources[datasource] = [record]
            else:
                record = data_sources[datasource][0]

            if product not in record['products']:
                record['products'].append(product)

        return data_sources

    def get_data_from_source(self) -> Iterable:
        """
        Gets the datasource/product data from the source.
        :return: Iterable, yields datasource, product, applicable_to
        """
        raise NotImplementedError()


class DatasourceCsv(DatasourceBase):
    """
    Import data from a CSV file, formatted Datasource,Product
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('DatasourceCsv: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--file', help='Path of the csv file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the datasource/product data from the source.
        :return: Iterable, yields datasource, product, applicable_to
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        with open(file) as f:
            lines = f.readlines()

        for detection in lines:
            parts = detection.split(',')
            data_source = parts[0].strip()
            product = parts[1].strip()
            yield data_source, product, None


class DatasourceExcel(DatasourceBase):
    """
    Import data sources from an Excel file, having a worksheet with two columns: Datasource and Product
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('DatasourceExcel: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--file', help='Path of the Excel file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the datasource/product data from the source.
        :return: Iterable, yields datasource, product, applicable_to
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        import openpyxl
        wb = openpyxl.load_workbook(filename=file, data_only=True)
        sheet = wb.worksheets[0]

        for rowNumber in range(2, sheet.max_row + 1):
            data_source = sheet.cell(row=rowNumber, column=1).value
            product = sheet.cell(row=rowNumber, column=2).value
            yield data_source, product, None


class DatasourceOssemBase(DatasourceBase):
    """
    Base class for importing datasource/product data that is based on OSSEM data
    For information about OSSEM see: https://github.com/OTRF/OSSEM-DM
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        self._log_source = None

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        raise NotImplementedError()

    def get_data_from_source(self) -> Iterable:
        """
        Gets the datasource/product data from the source.
        :return: Iterable, yields datasource, product, applicable_to
        """
        raise NotImplementedError()

    def _get_ossem_data(self):
        """
        Retrieves data from the OSSEM ATT&CK mapping
        """
        import pandas

        url = 'https://raw.githubusercontent.com/OTRF/OSSEM-DM/main/use-cases/mitre_attack/attack_events_mapping.csv'
        data = pandas.read_csv(url)
        data.where(data['Log Source'] == self._log_source, inplace=True)
        data.dropna(how='all', inplace=True)
        select = data[['Data Source', 'Component', 'EventID', 'Event Name', 'Filter in Log', 'Audit Category']]
        dict_result = select.to_dict(orient="records")
        return dict_result


class DatasourceDefenderEndpoints(DatasourceOssemBase):
    """
    Class for importing data source information for Microsoft Defender for Endpoints tables.
    Uses OSSEM to generate the overview.
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        self._log_source = 'Microsoft Defender for Endpoint'

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        pass

    def get_data_from_source(self) -> Iterable:
        """
        Gets the datasource/product data from the source.
        :return: Iterable, yields datasource, product, applicable_to
        """
        ossem_data = self._get_ossem_data()

        for record in ossem_data:
            action_types = json.loads(record['Filter in Log'].replace('\'', '\"'))
            if len(action_types) > 0:
                for action_type in action_types:
                    yield str(record['Component']).title(), f'{record ["Event Name"]}: {action_type["ActionType"]}', None
            else:
                yield str(record['Component']).title(), record['Event Name'], None


class DatasourceWindowsSysmon(DatasourceOssemBase):
    """
    Class for importing data source information for Sysmon.
    Uses your Sysmon config file and OSSEM to generate the overview.
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        if 'sysmon_config' not in self._parameters:
            raise Exception('DatasourceWindowsSysmon: "sysmon_config" parameter is required.')

        self._sysmon_config = parameters['sysmon_config']
        self._log_source = 'Microsoft-Windows-Sysmon'

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--sysmon_config', help='Path of the Sysmon config file.', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the datasource/product data from the source.
        :return: Iterable, yields datasource, product, applicable_to
        """
        ossem_data = self._get_ossem_data()
        sysmon_config = self._get_sysmon_config()

        for record in ossem_data:
            config_items = sysmon_config.findall(f'.//{record["Audit Category"]}')

            # If this is an event type with an onmatch == include attribute without child items this means
            # that nothing is being logged for this event type
            for config_item in config_items:
                if config_item.attrib['onmatch'] == "include" and len(config_item.getchildren()) == 0:
                    continue

            data_source = str(record['Component']).title()
            product = f'{record["EventID"]}: {record["Event Name"]}'
            yield data_source, product, None

    def _get_sysmon_config(self) -> Element:
        """
        Gets the Sysmon config from the filesystem
        """
        tree = ElementTree.parse(self._sysmon_config)
        root = tree.getroot()
        return root


class DatasourceWindowsSecurityAuditing(DatasourceOssemBase):
    """
    Class for importing data source information for Windows Security Auditing event logging.
    Uses the event id's logged in the last 30 days and OSSEM to generate the overview.
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        if 'app_id' not in self._parameters:
            raise Exception(f'DatasourceWindowsSecurityAuditing: "app_id" parameter is required.')

        if 'tenant_id' not in self._parameters:
            raise Exception(f'DatasourceWindowsSecurityAuditing: "tenant_id" parameter is required.')

        if 'workspace_id' not in self._parameters:
            raise Exception(f'DatasourceWindowsSecurityAuditing: "workspace_id" parameter is required.')

        self._app_id = self._parameters['app_id']
        self._tenant_id = self._parameters['tenant_id']
        self._workspace_id = self._parameters['workspace_id']
        self._secret = self._parameters.get('secret', None)
        self._endpoint = 'https://api.loganalytics.io'

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--app_id', help='Azure application id', required=True)
        parser.add_argument('--tenant_id', help='Azure tenant id', required=True)
        parser.add_argument('--workspace_id', help='Azure Sentinel workspace id', required=True)
        parser.add_argument('--secret', help='Azure client secret')

    def _connect_to_azure(self, endpoint: str) -> str:
        if self._secret:
            return Azure.connect_client_secret(self._app_id, self._tenant_id, endpoint, self._secret)
        else:
            return Azure.connect_device_flow(self._app_id, self._tenant_id, endpoint)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the datasource/product data from the source.
        :return: Iterable, yields datasource, product, applicable_to
        """
        access_token = self._connect_to_azure(self._endpoint)
        sentinel_data = self._get_sentinel_data(access_token)

        for record in sentinel_data:
            data_source = record[0].title()
            product = record[1]
            yield data_source, product, None

    def _get_sentinel_data(self, access_token: str) -> list:
        """
        Execute a query on Advanced Hunting to retrieve the datasource/product data
        :param access_token: JWT token to execute the request on the backend
        :return: Dictionary containing the results
        """
        url = f'{self._endpoint}/v1/workspaces/{self._workspace_id}/query'

        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': 'Bearer ' + access_token
        }

        query = '''
        let mapping_table = externaldata(DataSource:string,Component:string,Source:string,Relationship:string,Target:string, OSSEMID:string, EventID:int,EventName:string,EventPlatform:string, LogSource:string, FilterInLog:string,AuditCategory:string,AuditSubCategory:string, Channel:string, EnableCommands:string,GPOAuditPolicy:string)
        [
            h@"https://raw.githubusercontent.com/OTRF/OSSEM-DM/main/use-cases/mitre_attack/attack_events_mapping.csv"
        ]
        with(format="csv")
        | where LogSource == "Microsoft-Windows-Security-Auditing";
        // Get the event id's from the Windows Security Log and join this with the mapping table
        SecurityEvent
        | where TimeGenerated >= ago(30d)
        | where EventSourceName == "Microsoft-Windows-Security-Auditing"
        | distinct EventID
        | join kind = inner mapping_table on EventID
        // Get the distinct data sources
        | project  DataSource = Component, Product = strcat("Microsoft-Windows-Security-Auditing: ", tostring(EventID))
        | distinct DataSource, Product
        '''

        data = json.dumps({'query': query}).encode('utf-8')

        response = requests.post(url=url, headers=headers, data=data)

        if response.status_code != requests.codes['ok']:
            # Raise an exception to handle hitting API limits
            if response.status_code == requests.codes['too_many_requests']:
                raise ConnectionRefusedError('DatasourceWindowsSecurityAuditing: You have likely hit the API limit. ')
            response.raise_for_status()

        json_response = response.json()
        result = json_response['tables'][0]['rows']

        return result
