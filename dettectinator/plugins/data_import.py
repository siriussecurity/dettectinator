"""
Dettectinator - The Python library to your DeTT&CT YAML files.
Authors:
    Martijn Veken, Sirius Security
    Ruben Bouman, Sirius Security
License: GPL-3.0 License
"""

from argparse import ArgumentParser
from collections.abc import Iterable
import json
import os
import sys
import re
import requests
import urllib3

try:
    # When dettectinator is installed as python library
    from dettectinator.plugins.support.authentication import Azure, Tanium
except ModuleNotFoundError:
    # When dettectinator is not installed as python library
    sys.path.append(os.path.dirname(os.path.abspath(__file__).replace('plugins', '')))
    from plugins.support.authentication import Azure, Tanium


# Disable SSL certificate warnings for dev purposes:
urllib3.disable_warnings()


class ImportBase:
    """
    Base class for importing use-case/technique data
    """

    def __init__(self, parameters: dict) -> None:
        self._parameters = parameters

        self._re_include = self._parameters.get('re_include', None)
        self._re_exclude = self._parameters.get('re_exclude', None)

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        raise NotImplementedError()

    def get_attack_techniques(self, applicable_to: list, location_prefix: str) -> dict:
        """
        Retrieves use-case/technique data from a data source
        :param applicable_to: Systems that the detections are applicable to.
        :param location_prefix: Location of the detection, will be prepended to the detection name.
        :return: Dictionary, example: {'Detection A': {'applicable_to': ['all'], 'location_prefix': 'SIEM', 'techniques': ['T1055']}}
        """

        use_cases = {}

        for technique, use_case in self.get_data_from_source():
            # Exclude all detections that match the exclude pattern
            if self._re_exclude and not re.match(self._re_exclude, use_case) is None:
                continue

            # Include all detections that match the include pattern
            if self._re_include and re.match(self._re_include, use_case) is None:
                continue

            if use_case in use_cases.keys():
                use_cases[use_case]['techniques'].append(technique)
            else:
                use_cases[use_case] = {'applicable_to': applicable_to,
                                       'location_prefix': location_prefix,
                                       'techniques': [technique]}

        return use_cases

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        raise NotImplementedError()


class ImportCsv(ImportBase):
    """
    Import data from a CSV file, formatted TechniqueId,UseCase
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('ImportCsv: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--file', help='Path of the csv file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        with open(file) as f:
            lines = f.readlines()

        for detection in lines:
            parts = detection.split(',')
            technique = parts[0].strip()
            use_case = parts[1].strip()
            yield technique, use_case


class ImportExcel(ImportBase):
    """
    Import data from an Excel file, having a worksheet with two columns: TechniqueId and UseCase
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('ImportExcel: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--file', help='Path of the Excel file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        import openpyxl
        wb = openpyxl.load_workbook(filename=file, data_only=True)
        sheet = wb.worksheets[0]

        for rowNumber in range(2, sheet.max_row + 1):
            techniques = sheet.cell(row=rowNumber, column=1).value
            detection = sheet.cell(row=rowNumber, column=2).value
            for technique in techniques.split(','):
                yield technique.strip(), detection


class ImportAzureAuthBase(ImportBase):
    """
    Base class for import plugins that authenticate against Azure AD
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        if 'app_id' not in self._parameters:
            raise Exception(f'{self.__class__.__name__}: "app_id" parameter is required.')

        if 'tenant_id' not in self._parameters:
            raise Exception(f'{self.__class__.__name__}: "tenant_id" parameter is required.')

        self._app_id = self._parameters['app_id']
        self._tenant_id = self._parameters['tenant_id']
        self._secret = self._parameters.get('secret', None)

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--app_id', help='Azure application id', required=True)
        parser.add_argument('--tenant_id', help='Azure tenant id', required=True)
        parser.add_argument('--secret', help='Azure client secret')

    def get_data_from_source(self) -> Iterable:
        """
         Gets the use-case/technique data from the source.
         :return: Iterable, yields technique, detection
         """
        raise NotImplementedError()

    def _connect_to_azure(self, endpoint: str) -> str:
        if self._secret:
            return Azure.connect_client_secret(self._app_id, self._tenant_id, endpoint, self._secret)
        else:
            return Azure.connect_device_flow(self._app_id, self._tenant_id, endpoint)


class ImportSentinelAlertRules(ImportAzureAuthBase):
    """
    Import Analytics Rules from the Sentinel API
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        if 'subscription_id' not in self._parameters:
            raise Exception('ImportSentinelAlertRules: "subscription_id" parameter is required.')

        if 'resource_group' not in self._parameters:
            raise Exception('ImportSentinelAlertRules: "resource_group" parameter is required.')

        if 'workspace' not in self._parameters:
            raise Exception('ImportSentinelAlertRules: "workspace" parameter is required.')

        self._subscription_id = self._parameters['subscription_id']
        self._resource_group = self._parameters['resource_group']
        self._workspace = self._parameters['workspace']
        self._endpoint = 'https://management.azure.com'

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        ImportAzureAuthBase.set_plugin_params(parser)

        parser.add_argument('--subscription_id', help='Azure subscription id for Sentinel', required=True)
        parser.add_argument('--resource_group', help='Azure resource group for Sentinel', required=True)
        parser.add_argument('--workspace', help='Azure workspace for Sentinel', required=True)

    def get_data_from_source(self) -> Iterable:
        """
         Gets the use-case/technique data from the source.
         :return: Iterable, yields technique, detection
         """
        access_token = self._connect_to_azure(self._endpoint)
        sentinel_data = self._get_sentinel_data(access_token)

        for record in sentinel_data:
            properties = record['properties']

            if 'techniques' in properties and properties['techniques']:
                for technique in properties['techniques']:
                    use_case = properties['displayName']
                    yield technique, use_case

    def _get_sentinel_data(self, access_token: str) -> list:
        """
        Execute a query on Advanced Hunting to retrieve the use-case/technique data
        :param access_token: JWT token to execute the request on the backend
        :return: Dictionary containing the results
        """
        url = f'https://management.azure.com/subscriptions/{self._subscription_id}/resourceGroups/{self._resource_group}/' + \
              f'providers/Microsoft.OperationalInsights/workspaces/{self._workspace}/providers/Microsoft.SecurityInsights/' + \
              'alertRules?api-version=2022-07-01-preview'

        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': 'Bearer ' + access_token
        }
        response = requests.get(url=url, headers=headers)

        if response.status_code != requests.codes['ok']:
            # Raise an exception to handle hitting API limits
            if response.status_code == requests.codes['too_many_requests']:
                raise ConnectionRefusedError('ImportSentinelAlerts: You have likely hit the API limit. ')
            response.raise_for_status()

        json_response = response.json()
        result = json_response['value'] if 'value' in json_response else []

        return result


class ImportDefenderIdentityRules(ImportBase):
    """
    Import rules for Microsoft Defender for Identity from their Github webpage:
    https://github.com/MicrosoftDocs/ATADocs/tree/master/ATPDocs
    More info:
    https://learn.microsoft.com/en-us/defender-for-identity/alerts-overview
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        self.ATP_DOCS = ['https://raw.githubusercontent.com/MicrosoftDocs/ATADocs/master/ATPDocs/compromised-credentials-alerts.md',
                         'https://raw.githubusercontent.com/MicrosoftDocs/ATADocs/master/ATPDocs/domain-dominance-alerts.md',
                         'https://raw.githubusercontent.com/MicrosoftDocs/ATADocs/master/ATPDocs/exfiltration-alerts.md',
                         'https://raw.githubusercontent.com/MicrosoftDocs/ATADocs/master/ATPDocs/lateral-movement-alerts.md',
                         'https://raw.githubusercontent.com/MicrosoftDocs/ATADocs/master/ATPDocs/reconnaissance-alerts.md']

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        pass

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        for source_url in self.ATP_DOCS:
            resp = requests.get(source_url)
            body = resp.text

            # Remove comments from file, because it may contain commented detection rules.
            while '<!--' in body:
                body = body[0:body.find('<!--')] + body[body.find('-->')+3:]

            regex_title = re.compile('##\s(.*\s\(external\sID\s\d{4}\))')
            regex_tech = re.compile('\((T\d{4})\)')
            regex_subtech = re.compile('(T\d{4}\.\d{3})')

            current_detection = None
            for line in body.splitlines():
                title_match = regex_title.match(line)
                if title_match or current_detection is None:
                    if title_match:
                        current_detection = title_match.group(1)
                        continue
                else:
                    if 'MITRE attack technique' in line and 'N/A' in line:
                        current_detection = None
                    elif 'MITRE attack technique' in line:
                        tech_match = regex_tech.findall(line)
                        if tech_match:
                            for t in tech_match:
                                yield t, current_detection
                    elif 'MITRE attack sub-technique' in line and 'N/A' in line:
                        current_detection = None
                    elif 'MITRE attack sub-technique' in line:
                        subtech_match = regex_subtech.findall(line)
                        if subtech_match:
                            for t in subtech_match:
                                yield t, current_detection
                            current_detection = None


class ImportDefenderAlerts(ImportAzureAuthBase):
    """
    Import alerts and techniques from the Microsft Defender API.
    """
    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        self._endpoint = 'https://api.security.microsoft.com'

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        access_token = self._connect_to_azure(self._endpoint)
        defender_data = self._get_defender_data(access_token)

        for record in defender_data:
            technique = record['TechniqueId']
            use_case = record['Title'].strip()
            yield technique, use_case

    @staticmethod
    def _get_defender_data(access_token: str) -> dict:
        """
        Execute a query on Advanced Hunting to retrieve the use-case/technique data
        :param access_token: JWT token to execute the request on the backend
        :return: Dictionary containing the results
        """
        query = '''
        AlertInfo
        | mv-expand todynamic(AttackTechniques)
        | extend TechniqueId = extract(@'\((T.*)\)', 1, tostring(AttackTechniques))
        | where isnotempty(TechniqueId)
        | distinct TechniqueId, Title
        | order by TechniqueId
        '''

        url = 'https://api.security.microsoft.com/api/advancedhunting/run'
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': 'Bearer ' + access_token
        }

        data = json.dumps({'Query': query}).encode('utf-8')

        response = requests.post(url=url, headers=headers, data=data)

        if response.status_code != requests.codes['ok']:
            # Raise an exception to handle hitting API limits
            if response.status_code == requests.codes['too_many_requests']:
                raise ConnectionRefusedError('ImportDefenderAlerts: You have likely hit the API limit. ')
            response.raise_for_status()

        json_response = response.json()
        result = json_response['Results'] if 'Results' in json_response else {}

        return result


class ImportTaniumSignals(ImportBase):
    """
    Class for importing signals with ATT&CK technique mapping from Tanium.
    """
    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        if 'host' not in self._parameters:
            raise Exception('ImportTaniumSignals: "host" parameter is required.')
        if 'user' not in self._parameters:
            raise Exception('ImportTaniumSignals: "user" parameter is required.')
        if 'password' not in self._parameters:
            raise Exception('ImportTaniumSignals: "password" parameter is required.')
        if 'search_prefix' not in self._parameters:
            raise Exception('ImportTaniumSignals: "search_prefix" parameter is required.')

        self._host = self._parameters['host']
        self._user = self._parameters['user']
        self._password = self._parameters['password']
        self._search_prefix = self._parameters['search_prefix']
        self._LOGIN_URL = 'https://' + self._host + '/api/v2/session/login'
        self._INTEL_URL = 'https://' + self._host + '/plugin/products/detect3/api/v1/intels'

        self._session = Tanium.connect_http(self._user, self._password, self._LOGIN_URL)

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--host', help='Tanium host', required=True)
        parser.add_argument('--user', help='Tanium API username', required=True)
        parser.add_argument('--password', help='Tanium API password', required=True)
        parser.add_argument('--search_prefix', help='Search prefix')

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        tanium_data = self._get_all_signals()

        for signal in tanium_data:
            if 'mitreAttack' in signal.keys() and signal['mitreAttack']:
                signal_techniques = json.loads(signal['mitreAttack'])

                for t in signal_techniques['techniques']:
                    technique = t['id']
                    use_case = signal['name']
                    yield technique, use_case

    def _get_all_signals(self) -> dict:
        """
        Gets all signals (max 500 allowed by API) from Tanium.
        """
        headers = {'session': self._session, 'Content-Type': 'application/json'}
        params = {'limit': '500', 'name': self._search_prefix}
        r = requests.get(self._INTEL_URL, params=params, headers=headers, verify=False)
        if r.status_code == requests.codes.ok:
            return r.json()
        else:
            raise Exception(f'ImportTaniumSignals: get all signals failed: {r.text}')


class ImportElasticSecurityRules(ImportBase):
    """
    Class for importing Elastic Security rules with ATT&CK technique mapping.
    """
    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)

        if 'host' not in self._parameters:
            raise Exception('ImportElasticSecurityRules: "host" parameter is required.')
        if 'user' not in self._parameters:
            raise Exception('ImportElasticSecurityRules: "user" parameter is required.')
        if 'password' not in self._parameters:
            raise Exception('ImportElasticSecurityRules: "password" parameter is required.')

        self._host = self._parameters['host']
        self._user = self._parameters['user']
        self._password = self._parameters['password']
        self._filter = self._parameters['filter']
        self._FIND_URL = 'https://' + self._host + '/api/detection_engine/rules/_find'

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--host', help='Elastic Security host', required=True)
        parser.add_argument('--user', help='Elastic Security username', required=True)
        parser.add_argument('--password', help='Elastic Security password', required=True)
        parser.add_argument('--filter', help='Search filter, see Elastic documentation for more information')

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        rule_data = self._get_all_rules()

        for rule in rule_data['data']:
            if 'threat' in rule.keys():
                for threat in rule['threat']:
                    if threat['framework'] == 'MITRE ATT&CK' and 'technique' in threat.keys():
                        for tech in threat['technique']:
                            if 'subtechnique' in tech.keys() and len(tech['subtechnique']) > 0:
                                for subtech in tech['subtechnique']:
                                    technique = subtech['id']
                                    use_case = rule['name']
                                    yield technique, use_case
                            else:
                                technique = tech['id']
                                use_case = rule['name']
                                yield technique, use_case

    def _get_all_rules(self):
        headers = {'kbn-xsrf': 'dettect', 'Content-Type': 'application/json'}
        params = {'per_page': '10000'}
        if self._filter:
            params['filter'] = self._filter
        r = requests.get(self._FIND_URL, params=params, headers=headers, auth=(self._user, self._password), verify=False)
        if r.status_code == requests.codes.ok:
            return r.json()
        else:
            raise Exception(f'ImportElasticSecurityRules: get all rules failed: {r.text}')


class ImportSuricataRules(ImportBase):
    """
    Import data from a Suricata rules file. It expects a metadata meta-setting containing a field with the name
    mitre_technique_id containing the ATT&CK technique ID.

    https://suricata.readthedocs.io/en/latest/rules/meta.html#metadata

    Example (taken from https://rules.emergingthreats.net/open/suricata/rules/emerging-hunting.rules):
    alert http $HOME_NET any -> $EXTERNAL_NET any (msg:"ET HUNTING Possible Phishing - Form submitted to submit-form Form Hosting";
    flow:established,to_server; http.method; content:"POST"; http.host; content:"submit-form.com"; endswith; classtype:credential-theft;
    sid:2030707; rev:2; metadata:affected_product Web_Browsers, attack_target Client_Endpoint, created_at 2020_08_20, deployment Perimeter,
    former_category HUNTING, signature_severity Critical, tag Phishing, updated_at 2020_08_20, mitre_tactic_id TA0001, mitre_tactic_name Initial_Access,
    mitre_technique_id T1566, mitre_technique_name Phishing;)
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('ImportSuricateRules: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--file', help='Path of the Suricate rules file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        from suricataparser import parse_file
        rules = parse_file(file)

        for rule in rules:
            if rule.enabled:
                for option in rule.options:
                    if option.name == 'metadata':
                        meta_data = self._convert_metadata_list_to_dict(option.value.data)
                        if 'mitre_technique_id' in meta_data.keys():
                            yield meta_data['mitre_technique_id'], rule.msg

    def _convert_metadata_list_to_dict(self, meta_data: list) -> dict:
        """
        Converts a list with "key<space>value" into a dictionary.
        """
        meta_data_dict = {}
        for item in meta_data:
            splitted = item.split(' ')
            meta_data_dict[splitted[0]] = splitted[1]
        return meta_data_dict


class ImportSigmaRules(ImportBase):
    """
    Import data from a folder with Sigma rules.
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'folder' not in self._parameters:
            raise Exception('ImportSigmaRules: "folder" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--folder', help='Path of the folder with Sigma rules to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        folder = self._parameters['folder']

        if not os.path.isdir(folder):
            raise Exception(f'Folder does not exist: {folder}')

        from ruamel.yaml import YAML

        print(f'Reading data from "{folder}"')

        for root, _, files in os.walk(folder):
            for file in files:
                if file.endswith('.yaml') or file.endswith('.yml'):
                    filename = os.path.join(root, file)
                    yaml = YAML()
                    try:
                        with open(filename, 'r') as yaml_file:
                            yaml_content = yaml.load(yaml_file)
                    except Exception as e:
                        raise Exception(f'Failed loading YAML file "{filename}". Error: {str(e)}') from e

                    if 'tags' in yaml_content.keys():
                        for tag in yaml_content['tags']:
                            if tag.startswith('attack.t'):
                                yield tag[7:].upper(), yaml_content['title']


class ImportSplunkConfigSearches(ImportBase):
    """
    Import data from a Splunk config that contains saved searches (savedsearches.conf). It uses
    the action.correlationsearch.annotations attribute to get the mitre_attack techniques:

    action.correlationsearch.annotations = {"mitre_attack": ["T1560.001", "T1560"]}

    Searches that contain a action.correlationsearch.label and don't have disabled=1 are included.
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('ImportSplunkConfigSearches: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        parser.add_argument('--file', help='Path of the savedsearches config file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        import addonfactory_splunk_conf_parser_lib as splunk_conf_parser
        splunk_config = None
        with open(file, "r") as f:
            splunk_config = splunk_conf_parser.TABConfigParser()
            splunk_config.read_file(f)

        IGNORE_LIST = ['default']
        for section in splunk_config.sections():
            if splunk_config[section].name in IGNORE_LIST \
               or 'action.correlationsearch.label' not in splunk_config[section].keys() \
               or 'action.correlationsearch.annotations' not in splunk_config[section].keys() \
               or ('disabled' in splunk_config[section].keys() and splunk_config[section]['disabled'] == '1'):
                continue

            try:
                annotations = json.loads(splunk_config[section]['action.correlationsearch.annotations'])
            except Exception as e:
                print(f'Could not parse mitre_attack entry in action.correlationsearch.annotations ({str(e)}): {splunk_config[section].name}')
            else:
                if 'mitre_attack' in annotations.keys():
                    for technique in annotations['mitre_attack']:
                        yield technique, splunk_config[section].name
